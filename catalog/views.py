from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Q
from .models import Book, Section
from .forms import BookForm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse

def book_list(request):
    books = Book.objects.all()
    sections = Section.objects.all()
    
    # НОВОЕ: Получаем список всех существующих годов (исключая пустые), уникальные значения, сортируем по убыванию (новые сверху)
    years = Book.objects.exclude(publication_year__isnull=True).values_list('publication_year', flat=True).distinct().order_by('-publication_year')

    # Получаем параметры с формы
    query = request.GET.get('q') 
    section_id = request.GET.get('section')
    inv_num = request.GET.get('inv_num') 
    year = request.GET.get('year') # Выбранный год из списка
    sort_by = request.GET.get('sort_by') 

    # Фильтрация
    if query:
        books = books.filter(Q(title__icontains=query) | Q(author__icontains=query))
    if section_id:
        books = books.filter(section_id=section_id)
    if inv_num:
        books = books.filter(inventory_number__icontains=inv_num)
    if year: 
        books = books.filter(publication_year=year)
        
    # СОРТИРОВКА (теперь по новым годам по умолчанию)
    if sort_by == 'author':
        books = books.order_by('author')
    elif sort_by == 'title':
        books = books.order_by('title')
    elif sort_by == 'year_asc':
        books = books.order_by('publication_year') # Старые сначала
    else:
        # ПО УМОЛЧАНИЮ (если ничего не выбрано или выбрано year_desc):
        books = books.order_by('-publication_year', 'title')

    context = {
        'books': books,
        'sections': sections,
        'years': years, # Передаем собранные года в шаблон
    }
    return render(request, 'catalog/book_list.html', context)

def book_create(request):
    existing_book_id = None
    if request.method == 'POST':
        form = BookForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('book_list')
        else:
            # ПРОВЕРКА НА ДУБЛИКАТ: Если такой инвентарный номер уже есть
            if 'inventory_number' in form.errors:
                inv_num = request.POST.get('inventory_number')
                duplicate = Book.objects.filter(inventory_number=inv_num).first()
                if duplicate:
                    existing_book_id = duplicate.id # Передаем ID для кнопки "Пересохранить"
    else:
        form = BookForm()
    
    return render(request, 'catalog/book_form.html', {'form': form, 'existing_book_id': existing_book_id})

# НОВАЯ ФУНКЦИЯ ДЛЯ ПЕРЕСОХРАНЕНИЯ (РЕДАКТИРОВАНИЯ)
def book_update(request, pk):
    book = get_object_or_404(Book, pk=pk)
    if request.method == 'POST':
        form = BookForm(request.POST, instance=book)
        if form.is_valid():
            form.save()
            return redirect('book_list')
    else:
        form = BookForm(instance=book)
    
    return render(request, 'catalog/book_form.html', {'form': form, 'is_update': True})


def export_books_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="library_books.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Каталог книг'

    header_font = Font(name='Arial', bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin', color='BDC3C7'), right=Side(style='thin', color='BDC3C7'), top=Side(style='thin', color='BDC3C7'), bottom=Side(style='thin', color='BDC3C7'))

    # ДОБАВЛЕНО "Местонахождение" после года
    columns = ['Инв. №', 'Название', 'Автор', 'Раздел', 'Шифр', 'Год', 'Местонахождение', 'Стоимость']
    worksheet.append(columns)

    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Настраиваем ширину новой колонки (G)
    col_widths = {'A': 15, 'B': 45, 'C': 30, 'D': 25, 'E': 15, 'F': 10, 'G': 25, 'H': 15}
    for col, width in col_widths.items():
        worksheet.column_dimensions[col].width = width

    for idx, book in enumerate(Book.objects.all(), start=2):
        section_name = book.section.name if book.section else ""
        
        # ДОБАВЛЕНО: book.address
        row = [
            book.inventory_number, book.title, book.author, section_name, 
            book.cipher, book.publication_year, book.address, book.cost
        ]
        worksheet.append(row)
        
        for col_num in range(1, len(row) + 1):
            cell = worksheet.cell(row=idx, column=col_num)
            cell.border = thin_border
            
            # Центрируем 1, 5, 6 и 8 колонки
            if col_num in [1, 5, 6, 8]: 
                cell.alignment = center_align
            else: 
                cell.alignment = left_align
                
            if idx % 2 == 0:
                cell.fill = alt_row_fill

    worksheet.freeze_panes = 'A2'
    workbook.save(response)
    return response