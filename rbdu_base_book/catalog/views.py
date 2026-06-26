from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Q, Count
from .models import Book, Section, Address
from .forms import BookForm
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
import openpyxl
from django.http import Http404
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse

def reader_home(request):
    """Главная страница читателя: красивый поиск и плитка категорий"""
    sections = Section.objects.annotate(total=Count('book')).filter(total__gt=0)
    return render(request, 'catalog/reader_home.html', {'sections': sections})

def book_list(request):
    books = Book.objects.all()
    sections = Section.objects.all()
    
    years = Book.objects.exclude(publication_year__isnull=True).values_list('publication_year', flat=True).distinct().order_by('-publication_year')

    # Получаем параметры фильтрации
    query = request.GET.get('q') 
    section_id = request.GET.get('section')
    inv_num = request.GET.get('inv_num') 
    year = request.GET.get('year') 
    sort_by = request.GET.get('sort_by') 

    # [Блок фильтрации остается прежним]
    if query:
        books = books.filter(Q(title__icontains=query) | Q(author__icontains=query))
    if section_id:
        books = books.filter(section_id=section_id)
    if inv_num:
        books = books.filter(inventory_number__icontains=inv_num)
    if year: 
        books = books.filter(publication_year=year)
        
    if sort_by == 'author':
        books = books.order_by('author')
    elif sort_by == 'title':
        books = books.order_by('title')
    elif sort_by == 'year_asc':
        books = books.order_by('publication_year') 
    else:
        books = books.order_by('-publication_year', 'title')

    # --- НАЧАЛО БЛОКА ПАГИНАЦИИ ---
    paginator = Paginator(books, 20) # По 20 книг на страницу
    page_number = request.GET.get('page')
    
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1) # Если 'page' не число, даем первую страницу
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages) # Если страница за пределами — последнюю

    # Трюк: сохраняем GET-параметры фильтров для ссылок пагинации
    get_params = request.GET.copy()
    if 'page' in get_params:
        del get_params['page']
    url_params = get_params.urlencode() # Превратит фильтры в строку вида "q=привет&year=2024"
    # --- КОНЕЦ БЛОКА ПАГИНАЦИИ ---

    total_books = Book.objects.count()
    categories_stats = Section.objects.annotate(total=Count('book')).filter(total__gt=0)

    context = {
        'books': page_obj,          # Важно! Передаем объект страницы вместо всего QuerySet
        'sections': sections,
        'years': years,
        'total_books': total_books,      
        'categories_stats': categories_stats, 
        'url_params': url_params,   # Передаем параметры фильтров в шаблон
    }
    return render(request, 'catalog/book_list.html', context)

def book_detail(request, pk):  
    from django.shortcuts import get_object_or_404
    book = get_object_or_404(Book, pk=pk)
    return render(request, 'catalog/book_detail.html', {'book': book})

@login_required
def add_section(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        if name:
            Section.objects.create(name=name)
    return redirect(request.META.get('HTTP_REFERER', 'book_list'))


@login_required
def book_create(request):
    existing_book_id = None

    if request.method == 'POST':
        form = BookForm(request.POST)

        if form.is_valid():
            form.save()
            return redirect('book_list')

        print("FORM ERRORS:", form.errors)

        if 'inventory_number' in form.errors:
            inv_num = request.POST.get('inventory_number')
            duplicate = Book.objects.filter(
                inventory_number=inv_num
            ).first()

            if duplicate:
                existing_book_id = duplicate.id

    else:
        form = BookForm()

    return render(
        request,
        'catalog/book_form.html',
        {
            'form': form,
            'existing_book_id': existing_book_id
        }
    )

@login_required
def add_address(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        if name:
            Address.objects.get_or_create(name=name)
    return redirect(request.META.get('HTTP_REFERER', 'book_list'))

@login_required
def book_update(request, pk):
    book = get_object_or_404(Book, pk=pk)

    if request.method == 'POST':
        form = BookForm(request.POST, instance=book)

        if form.is_valid():
            form.save()
            return redirect('book_list')

        print("FORM ERRORS:", form.errors)

    else:
        form = BookForm(instance=book)

    return render(
        request,
        'catalog/book_form.html',
        {
            'form': form,
            'is_update': True,
            'book': book
        }
    )

@login_required
def book_delete(request, pk):
    book = get_object_or_404(Book, pk=pk)
    if request.method == 'POST':
        book.delete()
    return redirect('book_list')

@login_required  # Выгрузка доступна только авторизованным библиотекарям
def export_books_excel(request):
    # 1. Собираем те же фильтры, что используются на главной странице
    q = request.GET.get('q', '')
    section_id = request.GET.get('section', '')
    inv_num = request.GET.get('inv_num', '')
    year = request.GET.get('year', '')
    sort_by = request.GET.get('sort_by', 'year_desc')

    # 2. ВАЖНО: select_related('section') делает SQL JOIN и решает проблему WORKER TIMEOUT!
    books = Book.objects.select_related('section').all()

    # Применяем фильтрацию
    if q:
        books = books.filter(Q(title__icontains=q) | Q(author__icontains=q))
    if section_id:
        books = books.filter(section_id=section_id)
    if inv_num:
        books = books.filter(inventory_number__icontains=inv_num)
    if year:
        books = books.filter(publication_year=year)

    # Применяем сортировку
    if sort_by == 'year_asc':
        books = books.order_by('publication_year')
    elif sort_by == 'title':
        books = books.order_by('title')
    elif sort_by == 'author':
        books = books.order_by('author')
    else:
        books = books.order_by('-publication_year')

    # 3. Создаем Excel-файл в памяти
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отфильтрованный каталог"

    # Стилизуем заголовки
    headers = ["Инв. №", "Шифр", "Автор", "Название", "Раздел", "Год издания", "Местонахождение", "Цена"]
    ws.append(headers)

    # Заполняем данными (теперь это отработает мгновенно)
    for book in books:
        ws.append([
            book.inventory_number,
            book.cipher,
            book.author or "—",
            book.title,
            book.section.name if book.section else "—",  # Теперь это НЕ делает запрос в БД в цикле
            book.publication_year or "—",
            book.address or "—",
            book.cost or 0
        ])

    # Автоматически настраиваем ширину колонок для красоты
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 4. Формируем HTTP-ответ для скачивания файла
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="books_report.xlsx"'
    
    wb.save(response)
    return response